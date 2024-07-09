from flask import Flask, jsonify, render_template, request, send_file
from flask.views import MethodView
from flask_cors import CORS
from dotenv import load_dotenv
from openpyxl import Workbook, load_workbook
from pathlib import Path
from os import getenv
import json
from src.services.fetch_services import StockUpdater, SalesUpdate
from src.services.products_finder import ProductFinder
from src.services.product_services import ProductServices
from src.database.sql_operator import SQLOperator
from src.repository.pessoaRepository import PessoaRepository
from src.model.pessoa import Pessoa
from src.model.document import Document
from src.model.docitem import Docitem

from src.database.db_session_maker import Session_Maker


load_dotenv()
report_path = Path(getenv('STATIC') + 'report.json')
spreadsheet_path = Path(getenv('STATIC') + 'pedido.xlsx')


app = Flask(__name__)
CORS(app)


class HomePage(MethodView):
    def get(self):
        finder = ProductFinder()
        session_maker = Session_Maker()
        session = session_maker.create_session()
        pessoa_repo = PessoaRepository(session, Pessoa)
        reponse = pessoa_repo.find_all_suppliers()
        input(reponse)
        suppliers = finder.find_all_suppliers()
        groups = finder.find_all_prod_groups()
        families = finder.find_all_prod_family()
        response = {'fornecedores': suppliers, 'groups': groups, 'families': families}
        return jsonify(response)
        
class ListaCompras(MethodView):
    def get(self):
        data = request.args.get('data', False)
        if data:            
            parsed_data = json.loads(data)
            sql_operator = SQLOperator()
            product_services = ProductServices(sql_operator)
            filtered_products = product_services.join_groups_n_family(parsed_data)
            supplier_products = product_services.join_suppliers(parsed_data)
            all_products = filtered_products + supplier_products
            updated_products = StockUpdater.join_products_stock(all_products)
            result = SalesUpdate.join_product_sales(updated_products)
            
            with open(report_path, 'w+') as file:
                json.dump(result, file, indent=4)
                
            return jsonify({'ok': True}), 200
        else:
            return jsonify({'ok': False}), 404

class Informacoes(MethodView):
    def get(self):
        if report_path.exists():
            with open(report_path, 'r') as file:
                result = json.load(file)
            PATH = ('prod_list.html')
            input()
            return render_template(PATH, products=result)
        else:
            response = {'message': 'O relatorio não foi gerado, tente novamente'}
            return jsonify(response)

class CreateSpreadsheet(MethodView):
    def get(self):
        data = request.args.get('data', False)
        if data:
            parsed_data = json.loads(data)
            parsed_data.sort(key=lambda x: x['descricao'])
            wb = Workbook()
            sheet = wb.active
            sheet['A1'] = 'codigo'
            sheet['B1'] = 'descricao'
            sheet['C1'] = 'compra'
            sheet['D1'] = 'vl final'
            sheet['E1'] = 'vl sugestao'
            
            for index, item in enumerate(parsed_data, start=2):
                sheet[f'A{index}'] = item['codigo']
                sheet[f'B{index}'] = item['descricao']
                sheet[f'C{index}'] = item['compra']
                sheet[f'D{index}'] = ''
                sheet[f'E{index}'] = ''
            
            wb.save(spreadsheet_path)
            return jsonify({'ok': True}), 200
        else:
            return jsonify({'ok': False}), 404

class DownloadSpreadsheet(MethodView):
    def get(self):
        return send_file(spreadsheet_path)

class ProcessSheet(MethodView):
    def post(self):
        file = request.files.get('file')
        
        if file and file.filename.endswith('.xlsx'):
            workbook = load_workbook(file)
            sheet = workbook.active
            
            data = []
            headers = [cell.value for cell in sheet[1]]
            
            for row in sheet.iter_rows(min_row=2, values_only=True):
                row_data = {}
                for header, value in zip(headers, row):
                    row_data[header] = value
                data.append(row_data)
            
            return jsonify({'data': data}), 200
        else:
            return jsonify({'error': 'Arquivo inválido'}), 400

class CarregarPlanilha(MethodView):
    def get(self):
        return render_template('load_sheet.html')

# Register the URLs with the classes
app.add_url_rule('/', view_func=HomePage.as_view('index_page'))
app.add_url_rule('/lista_compras', view_func=ListaCompras.as_view('lista_compras'))
app.add_url_rule('/informacoes', view_func=Informacoes.as_view('informacoes'))
app.add_url_rule('/create_spreadsheet', view_func=CreateSpreadsheet.as_view('create_spreadsheet'))
app.add_url_rule('/donwload_spreadsheet', view_func=DownloadSpreadsheet.as_view('donwload_spreadsheet'))
app.add_url_rule('/process_sheet', view_func=ProcessSheet.as_view('process_sheet'), methods=['POST'])
app.add_url_rule('/carregar_planilha', view_func=CarregarPlanilha.as_view('carregar_planilha'))

if __name__ == '__main__':
    app.run('0.0.0.0', 5000, True)
